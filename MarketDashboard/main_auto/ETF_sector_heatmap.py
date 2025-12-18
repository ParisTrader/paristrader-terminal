"""
Sector ETF daily return heatmap (last 20 trading days)

- Automatically detect the column that contains ETF tickers from sector_etf_stock_list.xlsx
  (header can be Ticker/Symbol/Code, etc.)
- Download roughly two months of daily adjusted close data from yfinance (with a 3-month buffer)
- Compute daily percentage returns and keep the last 20 trading days
- Shape: rows = tickers, columns = dates with the newest on the left
- Color anchors (linear gradient, clipped outside the range):
    -10% = bright red (#FF0000)
     -1% = dark red (#8B0000)
      0% = neutral (#F5F5F5)
     +1% = dark green (#006400)
    +10% = bright green (#00FF00)
- Generate both HTML (sector_etf_heatmap.html) and Excel (sector_etf_heatmap.xlsx);
  Excel cells are filled with static colors that approximate the HTML version and show values as 0.00%.
"""

import math
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

import numpy as np
import pandas as pd
import yfinance as yf
from matplotlib.colors import LinearSegmentedColormap, to_hex

from update_volatility_360d import update_volatility_360d_column


BASE_DIR = Path(__file__).resolve().parent
VOL_COLUMN = "VOLATILITY_360D"


warnings.simplefilter("ignore", category=FutureWarning)


def detect_ticker_column(df: pd.DataFrame) -> str:
    """
    Automatically detect the column that contains ETF tickers from the Excel header.
    Prefer column names that include 'ticker'/'symbol'/'code' (case-insensitive).
    If nothing matches, fall back to the first object-dtype column.
    """
    if df.empty:
        raise ValueError("sector_etf_stock_list.xlsx 为空或未读取到数据。")

    candidates = [c for c in df.columns if isinstance(c, str)]
    if not candidates:
        raise ValueError("未找到可用的表头（列名）。")

    # 1) Candidate columns based on header keywords
    key_words = ("ticker", "symbol", "code")
    header_match_cols = []
    for col in candidates:
        lower = col.strip().lower()
        if any(k in lower for k in key_words):
            header_match_cols.append(col)

    # 2) Score columns based on how ticker-like the values look (ASCII + reasonable length + alnum/.-=)
    def is_ticker_like(s: str) -> bool:
        if not isinstance(s, str):
            return False
        s = s.strip()
        if not s:
            return False
        if any(ord(ch) > 127 for ch in s):  # Exclude non-ASCII strings (often Chinese names)
            return False
        if len(s) > 15:
            return False
        # Allowed characters: letters, digits, dot, dash, equals
        return all(ch.isalnum() or ch in ".-=" for ch in s)

    def column_score(col: str) -> float:
        series = df[col].astype(str).fillna("")
        if series.empty:
            return 0.0
        sample = series.head(200)  # enough samples for scoring
        vals = [v for v in sample if v and v.lower() != "nan"]
        if not vals:
            return 0.0
        like = sum(1 for v in vals if is_ticker_like(v))
        return like / max(1, len(vals))

    scored = [(col, column_score(col)) for col in candidates if pd.api.types.is_object_dtype(df[col])]
    # Priority: header keyword match plus high score
    scored.sort(key=lambda x: (x[0] in header_match_cols, x[1]), reverse=True)
    if scored and scored[0][1] > 0:  # Only accept when the score is positive
        return scored[0][0]

    # last resort: the first column
    return str(candidates[0])


def read_tickers_from_excel(path: str = "sector_etf_stock_list.xlsx", df: Optional[pd.DataFrame] = None) -> List[str]:
    if df is None:
        df = pd.read_excel(path)
    col = detect_ticker_column(df)
    raw = (
        df[col]
        .astype(str)
        .str.strip()
        .replace({"": np.nan})
        .dropna()
        .unique()
        .tolist()
    )

    # Keep only entries that look like tickers (avoid interpreting Chinese industry names as codes)
    def is_ticker_like(s: str) -> bool:
        if not isinstance(s, str):
            return False
        s = s.strip()
        if not s:
            return False
        if any(ord(ch) > 127 for ch in s):
            return False
        if len(s) > 15:
            return False
        return all(ch.isalnum() or ch in ".-=" for ch in s)

    tickers = [t for t in raw if is_ticker_like(t)]

    if not tickers:
        raise ValueError("未能从 Excel 中识别到任何 ETF 代码。请检查文件与列名，或确认含代码的列为 ASCII 代码。")
    return tickers


def read_column_b_mapping(path: str = "sector_etf_stock_list.xlsx", df: Optional[pd.DataFrame] = None) -> dict:
    """
    Read the Excel file and build a mapping from the ticker (column A or the detected ticker column)
    to the value in column B.
    Returns: dict[ticker] -> column B string.
    """
    if df is None:
        df = pd.read_excel(path)
    ticker_col = detect_ticker_column(df)
    
    # Fetch column B (index 1, the second column)
    if df.shape[1] < 2:
        return {}
    
    b_col_idx = 1  # Column B is the second column, index 1
    b_col_name = df.columns[b_col_idx]
    
    mapping = {}
    for idx, row in df.iterrows():
        ticker_raw = str(row.get(ticker_col, "")).strip()
        if not ticker_raw or ticker_raw.lower() == "nan":
            continue
        # Validate ticker candidates
        if any(ord(ch) > 127 for ch in ticker_raw) or len(ticker_raw) > 15:
            continue
        if not all(ch.isalnum() or ch in ".-=" for ch in ticker_raw):
            continue
        
        b_value = row.iloc[b_col_idx]
        if pd.notna(b_value):
            mapping[ticker_raw] = str(b_value).strip()
        else:
            mapping[ticker_raw] = ticker_raw  # Fall back to the ticker itself when column B is empty
    
    return mapping


def read_column_a_mapping(path: str = "sector_etf_stock_list.xlsx", df: Optional[pd.DataFrame] = None) -> dict:
    """
    Read the Excel file and build a mapping from ticker -> column A value.
    Returns: dict[ticker] -> column A string.
    """
    if df is None:
        df = pd.read_excel(path)
    ticker_col = detect_ticker_column(df)
    
    # Fetch column A (index 0, the first column)
    if df.shape[1] < 1:
        return {}
    
    a_col_idx = 0  # Column A is the first column, index 0
    
    mapping = {}
    for idx, row in df.iterrows():
        ticker_raw = str(row.get(ticker_col, "")).strip()
        if not ticker_raw or ticker_raw.lower() == "nan":
            continue
        # Validate ticker candidates
        if any(ord(ch) > 127 for ch in ticker_raw) or len(ticker_raw) > 15:
            continue
        if not all(ch.isalnum() or ch in ".-=" for ch in ticker_raw):
            continue
        
        a_value = row.iloc[a_col_idx]
        if pd.notna(a_value):
            mapping[ticker_raw] = str(a_value).strip()
        else:
            mapping[ticker_raw] = ""  # Return an empty string when column A is blank
    
    return mapping


def download_prices(tickers: List[str], period: str = "3mo") -> pd.DataFrame:
    """
    Download roughly 2-3 months of daily adjusted close prices.
    Returns a DataFrame indexed by date with tickers as columns.
    """
    # Prefer bulk download first
    data = yf.download(tickers, period=period, auto_adjust=True, progress=False, group_by="column")
    prices = pd.DataFrame()
    if not data.empty:
        if isinstance(data.columns, pd.MultiIndex):
            prices = data["Close"].copy()
        else:
            # Single-ticker scenario
            colname = "Close" if "Close" in data.columns else list(data.columns)[0]
            one = data[[colname]].copy()
            one.columns = [tickers[0]]
            prices = one

    # If bulk download fails or is empty, fetch tickers one by one and drop missing ones
    if prices.empty or prices.dropna(how="all").empty:
        series_list = []
        valid = []
        skipped = []
        for tk in tickers:
            try:
                tdf = yf.download(tk, period=period, auto_adjust=True, progress=False)
                if not tdf.empty and "Close" in tdf.columns and not tdf["Close"].dropna().empty:
                    s = tdf["Close"].copy()
                    s.name = tk
                    series_list.append(s)
                    valid.append(tk)
                else:
                    skipped.append(tk)
            except Exception:
                skipped.append(tk)
        if not series_list:
            raise RuntimeError("从 yfinance 未获取到任何价格数据。请检查网络或代码列表。")
        prices = pd.concat(series_list, axis=1).sort_index()
        if skipped:
            print(f"以下代码下载失败或无数据，已跳过：{skipped}")
    return prices


def compute_daily_returns(prices: pd.DataFrame) -> pd.DataFrame:
    """
    Compute daily percentage returns.
    """
    returns_pct = prices.pct_change() * 100.0
    return returns_pct


def latest_20_days_transposed(returns_pct: pd.DataFrame) -> pd.DataFrame:
    """
    - Select the latest 20 trading days
    - Rows = ETF tickers, columns = date strings; columns ordered from newest (left) to oldest (right)
    """
    # Keep some buffer, drop rows that are all NaN, then take the latest 20 days
    last = returns_pct.tail(40).dropna(how="all").tail(20).copy()
    # Format the index as date strings
    last.index = last.index.strftime("%Y-%m-%d")
    df = last.T
    # Sort columns from newest to oldest so the freshest date is on the left
    cols_sorted = sorted(df.columns, reverse=True)
    df = df[cols_sorted]
    return df


def prepare_vol_series(meta_df: pd.DataFrame, ticker_col: str, index: pd.Index) -> pd.Series:
    meta_copy = meta_df.copy()
    meta_copy[ticker_col] = meta_copy[ticker_col].astype(str).str.strip()
    if VOL_COLUMN not in meta_copy.columns:
        return pd.Series(np.nan, index=index, dtype=float)
    vol_series = (
        meta_copy.set_index(meta_copy[ticker_col])[VOL_COLUMN]
        .pipe(pd.to_numeric, errors="coerce")
        .reindex(index)
    )
    return vol_series


def prepare_mu_series(meta_df: pd.DataFrame, ticker_col: str, index: pd.Index) -> pd.Series:
    """从元数据 DataFrame 中提取 MU_1D 列（日均收益 μ，百分数）。"""
    meta_copy = meta_df.copy()
    meta_copy[ticker_col] = meta_copy[ticker_col].astype(str).str.strip()
    if "MU_1D" not in meta_copy.columns:
        return pd.Series(np.nan, index=index, dtype=float)
    mu_series = (
        meta_copy.set_index(meta_copy[ticker_col])["MU_1D"]
        .pipe(pd.to_numeric, errors="coerce")
        .reindex(index)
    )
    return mu_series


def compute_zs_live(df_pct: pd.DataFrame, vol_series: pd.Series) -> pd.Series:
    """
    Compute ZS Live from the latest daily return and VOLATILITY_360D.
    VOLATILITY_360D is the 1-year annualized volatility estimated from the last 252 trading days.
    ZS Live = pct_today * 16.0 / VOLATILITY_360D
    Note: ZS Live still uses 0-mean approximation, while ZS 5D uses explicit μ correction.
    """
    if df_pct.empty:
        return pd.Series(dtype=float)

    pct_today = df_pct.iloc[:, 0].astype(float)
    aligned_vol = vol_series.reindex(df_pct.index)
    zs_live = pd.Series(np.nan, index=df_pct.index, dtype=float)
    mask_valid = (aligned_vol > 0) & pct_today.notna()
    zs_live[mask_valid] = pct_today[mask_valid] * 16.0 / aligned_vol[mask_valid]
    zs_live = zs_live.replace([np.inf, -np.inf], np.nan)
    return zs_live


def compute_zs_5d(df_pct: pd.DataFrame, vol_series: pd.Series, mu_series: pd.Series) -> pd.Series:
    """
    Compute ZS 5D using the 5-day z-score definition with explicit μ correction.
    Formula: Z_5d = (R_5d - 5μ) / (σ_1d * sqrt(5))
    where:
    - R_5d: 5-day cumulative return (decimal)
    - μ: 1-day average return (decimal, from MU_1D in percent)
    - σ_1d: 1-day volatility (decimal, derived from annualized volatility)
    The volatility series (VOLATILITY_360D) is the 1-year annualized volatility estimated from the last 252 trading days.
    """
    if df_pct.empty or df_pct.shape[1] < 5:
        return pd.Series(np.nan, index=df_pct.index, dtype=float)

    # 1) 最近 5 日的百分比涨跌幅（df_view 最左 5 列）
    last5_pct = df_pct.iloc[:, :5].astype(float)         # shape: (n_ticker, 5)

    # 2) 转为小数收益
    last5_dec = last5_pct / 100.0                         # r_t

    # 3) 5 日累计收益 R_5d = Π(1 + r_t) - 1
    R_5d = (1.0 + last5_dec).prod(axis=1) - 1.0           # pd.Series, index = df_view.index

    # 4) 日波动率 σ_1d（小数）：由年化波 VOLATILITY_360D（百分数）反推
    sigma_1d = vol_series.reindex(df_pct.index) / (16.0 * 100.0)  # VOL/16/100

    # 5) 日均收益 μ（小数）：MU_1D 从百分数转为小数
    mu_daily_dec = mu_series.reindex(df_pct.index) / 100.0

    # 6) ZS 5D 按照 Z = (R_5d - 5μ) / (σ_1d * sqrt(5)) 计算
    zs_5d = pd.Series(np.nan, index=df_pct.index)
    mask_valid_5d = (
        sigma_1d.notna() & (sigma_1d > 0) &
        R_5d.notna() &
        mu_daily_dec.notna()
    )

    zs_5d[mask_valid_5d] = (
        R_5d[mask_valid_5d] - 5.0 * mu_daily_dec[mask_valid_5d]
    ) / (sigma_1d[mask_valid_5d] * math.sqrt(5.0))

    # 7) 清理无穷值
    zs_5d = zs_5d.replace([np.inf, -np.inf], np.nan)
    return zs_5d


def build_colormap() -> LinearSegmentedColormap:
    """
    Custom 5-anchor color map (tuned deep red/green shades):
      -10%: #FF0000 (sharp red)
       -1%: #C62828 (aesthetic dark red, Material Design Red 800)
        0%: #F5F5F5 (near-neutral)
      +1%: #0FA84C (brighter dark green starting point)
      +10%: #00FF00 (sharp green)
    Linearly interpolate across anchors and later clamp with vmin=-10, vmax=+10.
    """
    # Linearly project [-10, 10] to [0, 1]
    def pos(v: float) -> float:
        return (v + 10.0) / 20.0

    anchor_points = [
        (pos(-10.0), "#FF0000"),
        (pos(-1.0), "#C62828"),  # more visually pleasing dark red
        (pos(0.0), "#F5F5F5"),
        (pos(1.0), "#0FA84C"),  # brighter dark green anchor
        (pos(10.0), "#00FF00"),
    ]
    cmap = LinearSegmentedColormap.from_list("sector_returns_5anchors", anchor_points, N=256)
    return cmap


def to_html_heatmap(
    df_pct: pd.DataFrame,
    html_path: str,
    ticker_to_colb: dict = None,
    ticker_to_cola: dict = None,
    zs_live: Optional[pd.Series] = None,
    zs_5d: Optional[pd.Series] = None,
) -> None:
    """
    Render the percentage DataFrame (rows = ETF tickers, columns = dates, values in %) as an HTML heatmap.
    - Latest date is in the leftmost column
    - Each cell shows a signed percentage with two decimals
    - Uses the custom 5-anchor gradient
    - ticker_to_colb: mapping from ticker -> column B label for the first column
    - ticker_to_cola: mapping from ticker -> column A label for the second column
    """
    cmap = build_colormap()
    if ticker_to_colb is None:
        ticker_to_colb = {}
    if ticker_to_cola is None:
        ticker_to_cola = {}
    # Use custom inline styles to control the header cell layout and caption style
    html = []
    html.append("<html><head><meta charset='utf-8'>")
    html.append("<title>Sector ETF Daily Return Heatmap (Last 20 Trading Days)</title>")
    html.append("""<style>
        body { margin: 0; padding: 16px 24px 32px 24px; background: #ffffff; min-width: 1600px; min-height: 950px; }
        table { border-collapse: collapse; font-family: Arial; font-size: 13px; table-layout: fixed; min-width: 1600px; }
        th, td { border: 1px solid #ddd; padding: 6px 8px; text-align: center; overflow: hidden; white-space: nowrap; }
        caption { caption-side: top; font-weight: 800; font-size: 20px; line-height: 1.6; margin: 12px 0 14px 0; }
        th { background: #f0f0f0; }
        /* Wider first column */
        th:first-child, td:first-child { width: 160px; min-width: 140px; }
        /* Second column (industry nickname) slightly narrower but wider than data columns */
        th:nth-child(2), td:nth-child(2) { width: 96px; min-width: 90px; }
        /* Third column (ZS Live) width, identical to data columns */
        th:nth-child(3), td:nth-child(3) { width: 72px; min-width: 72px; }
        /* Fourth column (ZS 5D) width, identical to data columns */
        th:nth-child(4), td:nth-child(4) { width: 72px; min-width: 72px; }
        /* Uniform width for the remaining data columns */
        th:not(:first-child):not(:nth-child(2)):not(:nth-child(3)):not(:nth-child(4)),
        td:not(:first-child):not(:nth-child(2)):not(:nth-child(3)):not(:nth-child(4)) { width: 72px; min-width: 72px; }
        .corner { position: relative; height: 44px; }
        /* Diagonal separator from bottom-left to top-right */
        .corner::before { content: ""; position: absolute; left: 0; top: 0; width: 100%; height: 100%;
            background: linear-gradient(to top right, transparent calc(50% - 0.5px), #999 calc(50% - 0.5px), #999 calc(50% + 0.5px), transparent calc(50% + 0.5px));
            pointer-events: none; z-index: 1; }
        .corner .date { position: absolute; right: 6px; top: 4px; font-size: 14px; line-height: 1; font-weight: bold; z-index: 2; }
        .corner .ticker { position: absolute; left: 6px; bottom: 4px; font-size: 14px; line-height: 1; font-weight: bold; z-index: 2; }
    </style></head><body>""")
    html.append("<table>")
    html.append("<caption>Sector ETF Daily Return Heatmap (Last 20 Trading Days)</caption>")
    # header
    html.append("<thead><tr>")
    # A1 corner: show Date in the upper-right and Ticker in the lower-left
    html.append("<th><div class='corner'><span class='date'>Date</span><span class='ticker'>Ticker</span></div></th>")
    # Second column: industry nickname
    html.append("<th>行業</th>")
    # Third column: ZS Live
    html.append("<th>ZS Live</th>")
    html.append("<th>ZS 5D</th>")
    for col in df_pct.columns:
        html.append(f"<th>{col}</th>")
    html.append("</tr></thead>")

    # body with colored cells
    html.append("<tbody>")
    for idx, row in df_pct.iterrows():
        # First column shows column B label if available; otherwise fall back to the ticker
        ticker_key = str(idx)
        first_col_text = ticker_to_colb.get(ticker_key, ticker_key)
        # Second column shows column A label (industry nickname)
        second_col_text = ticker_to_cola.get(ticker_key, "")
        html.append(f"<tr><td>{first_col_text}</td><td>{second_col_text}</td>")
        # Third column: ZS Live
        val_live = None
        if zs_live is not None and ticker_key in zs_live.index:
            val_live = zs_live.loc[ticker_key]
        if val_live is None or pd.isna(val_live):
            html.append("<td></td>")
        else:
            color_live = color_for_zs_live(float(val_live))
            style_attr = f" style='background-color:{color_live};'" if color_live else ""
            html.append(f"<td{style_attr}>{float(val_live):+.2f}</td>")

        # Fourth column: ZS 5D
        val_5d = None
        if zs_5d is not None and ticker_key in zs_5d.index:
            val_5d = zs_5d.loc[ticker_key]
        if val_5d is None or pd.isna(val_5d):
            html.append("<td></td>")
        else:
            color_5d = color_for_zs_live(float(val_5d))
            style_attr_5d = f" style='background-color:{color_5d};'" if color_5d else ""
            html.append(f"<td{style_attr_5d}>{float(val_5d):+.2f}</td>")
        for val in row.values:
            if pd.isna(val):
                html.append("<td></td>")
            else:
                color = color_from_cmap_percent(float(val), cmap)
                text = f"{val:+.2f}%"
                html.append(f"<td style='background-color:{color};'>{text}</td>")
        html.append("</tr>")
    html.append("</tbody></table></body></html>")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))


def color_from_cmap_percent(val_pct: Optional[float], cmap: LinearSegmentedColormap) -> Optional[str]:
    """
    Map a percentage value (%) to a hex color (e.g., '#AABBCC').
    Return None for None/NaN and clamp values outside [-10, 10].
    """
    if val_pct is None or (isinstance(val_pct, float) and np.isnan(val_pct)):
        return None
    v = max(-10.0, min(10.0, float(val_pct)))
    pos = (v + 10.0) / 20.0
    rgba = cmap(pos)
    return to_hex(rgba, keep_alpha=False)


def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    hex_color = hex_color.lstrip("#")
    if len(hex_color) != 6:
        raise ValueError(f"Invalid hex color: {hex_color}")
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    return "#{:02X}{:02X}{:02X}".format(*rgb)


def color_for_zs_live(value: Optional[float], max_abs: float = 1.0) -> Optional[str]:
    """
    Return a gradient color based on the ZS Live value.
    - Near zero stays uncolored
    - Positive values are green, negative values are red
    - Reaches the most vivid color once |value| hits max_abs
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if max_abs <= 0:
        return None
    val = float(value)
    intensity = min(abs(val) / max_abs, 1.0)
    if intensity == 0:
        return None
    base_color = "#0FA84C" if val > 0 else "#C62828"
    rgb_base = _hex_to_rgb(base_color)
    rgb_white = (255, 255, 255)
    blended = tuple(
        int(round(rgb_white[i] + (rgb_base[i] - rgb_white[i]) * intensity))
        for i in range(3)
    )
    return _rgb_to_hex(blended)


def to_excel_with_fills(
    df_pct: pd.DataFrame,
    xlsx_path: str,
    sheet_name: str = "Heatmap",
    ticker_to_colb: dict = None,
    ticker_to_cola: dict = None,
    zs_live: Optional[pd.Series] = None,
    zs_5d: Optional[pd.Series] = None,
) -> None:
    """
    Export the percentage DataFrame to Excel:
    - Write numeric values (as fraction = value/100) with number format '0.00%'
    - Fill each cell with static colors that resemble the HTML heatmap (no conditional formatting)
    - ticker_to_colb: mapping from ticker -> column B text for the first column
    - ticker_to_cola: mapping from ticker -> column A text for the second column
    """
    if ticker_to_colb is None:
        ticker_to_colb = {}
    if ticker_to_cola is None:
        ticker_to_cola = {}
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment

    df_fraction = df_pct.copy() / 100.0
    df_for_excel = df_fraction.copy()
    first_col_values = [ticker_to_colb.get(str(idx), str(idx)) for idx in df_for_excel.index]
    df_for_excel.insert(0, "Ticker", first_col_values)
    col_a_values = [ticker_to_cola.get(str(idx), "") for idx in df_pct.index]
    df_for_excel.insert(1, "行业昵称", col_a_values)
    if zs_live is not None:
        zs_aligned = pd.to_numeric(zs_live.reindex(df_pct.index), errors="coerce")
    else:
        zs_aligned = pd.Series(np.nan, index=df_pct.index, dtype=float)
    df_for_excel.insert(2, "ZS Live", zs_aligned.tolist())
    if zs_5d is not None:
        zs5_aligned = pd.to_numeric(zs_5d.reindex(df_pct.index), errors="coerce")
    else:
        zs5_aligned = pd.Series(np.nan, index=df_pct.index, dtype=float)
    df_for_excel.insert(3, "ZS 5D", zs5_aligned.tolist())
    df_for_excel.reset_index(drop=True, inplace=True)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_for_excel.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.book.save(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

    # Format column 3 (ZS Live) as numbers
    for r in range(2, max_row + 1):
        cell = ws.cell(r, 3)
        val = cell.value
        if val is None:
            continue
        try:
            cell.value = float(val)
        except Exception:
            cell.value = None
            continue
        cell.number_format = "0.00;-0.00"
        color_z = color_for_zs_live(cell.value)
        if color_z:
            cell.fill = PatternFill(
                start_color=color_z.replace("#", ""),
                end_color=color_z.replace("#", ""),
                fill_type="solid",
            )

    # Format column 4 (ZS 5D) as numbers
    for r in range(2, max_row + 1):
        cell = ws.cell(r, 4)
        val = cell.value
        if val is None:
            continue
        try:
            cell.value = float(val)
        except Exception:
            cell.value = None
            continue
        cell.number_format = "0.00;-0.00"
        color_z5 = color_for_zs_live(cell.value)
        if color_z5:
            cell.fill = PatternFill(
                start_color=color_z5.replace("#", ""),
                end_color=color_z5.replace("#", ""),
                fill_type="solid",
            )

    # Set percentage format for date columns (starting from column 5)
    for r in range(2, max_row + 1):
        for c in range(5, max_col + 1):
            ws.cell(r, c).number_format = "0.00%"

    # Fill date columns with colors
    cmap = build_colormap()
    for r in range(2, max_row + 1):
        for c in range(5, max_col + 1):
            v_fraction = ws.cell(r, c).value
            if v_fraction is None:
                continue
            try:
                v_pct = float(v_fraction) * 100.0
            except Exception:
                continue
            hex_color = color_from_cmap_percent(v_pct, cmap)
            if hex_color is None:
                continue
            fill = PatternFill(
                start_color=hex_color.replace("#", ""),
                end_color=hex_color.replace("#", ""),
                fill_type="solid",
            )
            ws.cell(r, c).fill = fill

    from openpyxl.utils import get_column_letter
    first_col_width = 22
    second_col_width = 14
    data_col_width = 12
    third_col_width = data_col_width
    fourth_col_width = data_col_width
    if ws.max_column >= 1:
        ws.column_dimensions[get_column_letter(1)].width = first_col_width
    if ws.max_column >= 2:
        ws.column_dimensions[get_column_letter(2)].width = second_col_width
    if ws.max_column >= 3:
        ws.column_dimensions[get_column_letter(3)].width = third_col_width
    if ws.max_column >= 4:
        ws.column_dimensions[get_column_letter(4)].width = fourth_col_width
    for c in range(5, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = data_col_width

    from openpyxl.styles import Font, Border, Side
    a1 = ws.cell(1, 1)
    a1.value = "              Date\nTicker"
    a1.alignment = Alignment(horizontal="left", vertical="distributed", wrap_text=True)
    a1.font = Font(name="Consolas", bold=True)
    ws.row_dimensions[1].height = 28
    a1.border = Border(diagonal=Side(style="thin", color="999999"), diagonalDown=True)

    wb.save(xlsx_path)


def generate_etf_sector_heatmap(output_html_path: str, output_excel_path: Optional[str] = None) -> str:
    """
    Generate the ETF sector heatmap report (HTML + Excel) and return the HTML path.
    """
    output_html = Path(output_html_path).resolve()
    output_html.parent.mkdir(parents=True, exist_ok=True)

    if output_excel_path is None:
        output_excel = output_html.with_suffix(".xlsx")
    else:
        output_excel = Path(output_excel_path).resolve()
        output_excel.parent.mkdir(parents=True, exist_ok=True)

    excel_path = BASE_DIR / "sector_etf_stock_list.xlsx"
    excel_path_str = str(excel_path)

    # 0) Refresh the 1-year annualized volatility column first (based on last 252 trading days)
    print("\n====== 阶段 1/2：波动率计算（VOLATILITY_360D 列刷新，基于近 252 个交易日）开始 ======", flush=True)
    update_volatility_360d_column(excel_path_str)
    print("====== 阶段 1/2：波动率计算完成 ======\n", flush=True)

    print("====== 阶段 2/2：Heatmap 生成中 ======", flush=True)

    meta_df = pd.read_excel(excel_path)
    ticker_col = detect_ticker_column(meta_df)

    # 1) Read the ETF column
    tickers = read_tickers_from_excel(excel_path_str, df=meta_df)
    print(f"读取到 {len(tickers)} 个 ETF 代码：{tickers}", flush=True)

    # 2) Download prices (~2 months, extra buffer by requesting 3mo)
    print("正在从 yfinance 下载数据（period='3mo'）...", flush=True)
    prices = download_prices(tickers, period="3mo")

    # 3) Compute daily percentage returns
    returns_pct = compute_daily_returns(prices)

    # 4) Keep the latest 20 trading days and transpose so the newest date is on the left
    df_view = latest_20_days_transposed(returns_pct)

    # 4.1) Load column B and column A mappings
    ticker_to_colb = read_column_b_mapping(excel_path_str, df=meta_df)
    ticker_to_cola = read_column_a_mapping(excel_path_str, df=meta_df)

    # 4.2) Calculate ZS Live and ZS 5D
    vol_series = prepare_vol_series(meta_df, ticker_col, df_view.index)
    mu_series = prepare_mu_series(meta_df, ticker_col, df_view.index)
    zs_live = compute_zs_live(df_view, vol_series)
    zs_5d = compute_zs_5d(df_view, vol_series, mu_series)

    # 5) Output HTML
    html_out = str(output_html)
    to_html_heatmap(
        df_view,
        html_out,
        ticker_to_colb=ticker_to_colb,
        ticker_to_cola=ticker_to_cola,
        zs_live=zs_live,
        zs_5d=zs_5d,
    )
    print(f"HTML 报告已生成：{html_out}", flush=True)

    # 6) Output Excel (static fills approximating the HTML colors)
    xlsx_out = str(output_excel)
    to_excel_with_fills(
        df_view,
        xlsx_out,
        sheet_name="Heatmap",
        ticker_to_colb=ticker_to_colb,
        ticker_to_cola=ticker_to_cola,
        zs_live=zs_live,
        zs_5d=zs_5d,
    )
    print(f"Excel 文件已生成：{xlsx_out}", flush=True)

    return html_out


def main() -> None:
    default_html = BASE_DIR / "sector_etf_heatmap.html"
    default_excel = BASE_DIR / "sector_etf_heatmap.xlsx"
    generate_etf_sector_heatmap(str(default_html), str(default_excel))


if __name__ == "__main__":
    main()