# -*- coding: utf-8 -*-
"""
CBBC HSI Notional 計算器 - V5 (診斷修正版)
重點修正：
1. 在解析前「強制移除所有空白行」，解決標題下方的空行導致讀取失敗的問題。
2. 將 on_bad_lines 改為 'warn'，如果資料被丟棄會顯示原因。
3. 增加資料預覽功能，確保篩選前資料已正確載入。
"""

import pandas as pd
import datetime as dt
import os
import requests
import yfinance as yf
import warnings
import time
from io import StringIO
from pathlib import Path
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from get_usdhkd import get_usdhkd_rate

warnings.simplefilter("ignore", FutureWarning)

# ==================== 路徑設定 ====================
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_ROOT = BASE_DIR / "cbbc_dataset"
DEFAULT_OUTPUT_DIR = BASE_DIR / "output"
# 可透過環境變數 CBBC_DATA_ROOT / CBBC_OUTPUT_DIR 覆寫上述預設路徑

ROOT_DIR = Path(os.environ.get("CBBC_DATA_ROOT", DEFAULT_DATA_ROOT))
raw_dir = ROOT_DIR / "raw"
spot_dir = Path(os.environ.get("CBBC_OUTPUT_DIR", DEFAULT_OUTPUT_DIR))

for path in (raw_dir, spot_dir):
    path.mkdir(parents=True, exist_ok=True)

strike_percent = ['10%', '9%', '8%', '7%', '6%', '5%', '4%', '3%', '2%', '1%', 'Spot',
                  '-1%', '-2%', '-3%', '-4%', '-5%', '-6%', '-7%', '-8%', '-9%', '-10%']

MAX_NETWORK_RETRIES = 5

def download_and_parse_v5():
    url = "https://www.hkex.com.hk/eng/cbbc/search/cbbcfulllist.csv"
    headers = {"User-Agent": "Mozilla/5.0"}

    print("1. [下載] 正在下載 HKEX CBBC 全表...")
    try:
        r = None
        for attempt in range(1, MAX_NETWORK_RETRIES + 1):
            try:
                r = requests.get(url, headers=headers, timeout=30)
                r.raise_for_status()
                break
            except requests.RequestException as err:
                if attempt == MAX_NETWORK_RETRIES:
                    print(f"   -> 下載失敗 ({attempt}/{MAX_NETWORK_RETRIES})：{err}")
                    raise
                wait_s = min(5, 2 ** attempt)
                print(f"   -> 下載失敗 ({err})，{wait_s}s 後重試 ({attempt}/{MAX_NETWORK_RETRIES})")
                time.sleep(wait_s)

        # 解碼
        try:
            text = r.content.decode('utf-16')
        except:
            print("   -> UTF-16 解碼失敗，嘗試 UTF-8")
            text = r.content.decode('utf-8', errors='ignore')

        # --- 步驟 A: 定位標題列 ---
        lines = text.split('\n')
        start_row = -1
        for i, line in enumerate(lines[:50]):
            # 尋找特徵：同時包含 Code 和 Bull
            if "Code" in line and "Bull" in line:
                start_row = i
                break

        if start_row == -1:
            raise Exception("找不到標題列！")

        print(f"   -> 在第 {start_row + 1} 行找到標題列")

        # --- 步驟 B: 強制清理空白行 (解決您的疑慮) ---
        # 取出標題列之後的所有行
        raw_lines = lines[start_row:]
        # 過濾掉只有空白符號或完全空的行
        clean_lines = [line for line in raw_lines if line.strip()]

        print(f"   -> 清理前行數: {len(raw_lines)}, 清理後行數: {len(clean_lines)}")

        # 重新組合成字串
        clean_text = '\n'.join(clean_lines)

        # --- 步驟 C: 讀取並顯示警告 (warn) ---
        # 使用 on_bad_lines='warn'，如果資料行格式不對，會直接印出錯誤訊息
        df = pd.read_csv(
            StringIO(clean_text),
            sep=None,
            engine="python",
            on_bad_lines='warn'
        )

        # 去除欄位空白
        df.columns = df.columns.str.strip()

        # --- 步驟 D: 檢查資料是否讀入 ---
        print(f"   -> 讀取完成，資料維度 (Shape): {df.shape}")
        if df.empty:
            raise Exception("讀取後 DataFrame 為空！可能是分隔符號錯誤或 Bad Lines 全部被丟棄。")

        # 簡單預覽
        print("   -> 資料預覽 (前 1 筆):")
        print(df.iloc[0].to_dict())

        # --- 步驟 E: 欄位映射（按實際資料調整） ---
        # 觀察 df 頭幾行可知：
        # - "Issuer" 列存的是標的代碼（例如 SPX、NDX、HSI 等）
        # - "UL" 列存的是 Bull / Bear 方向
        # 因此這裡直接將：
        #   Issuer -> UL（標的）
        #   UL     -> Bull/Bear（牛熊方向）

        if 'Issuer' not in df.columns or 'UL' not in df.columns:
            raise KeyError(f"找不到 Issuer / UL 欄位，現有欄位: {list(df.columns)}")

        ul_col = 'Issuer'   # 標的列
        bb_col = 'UL'       # Bull / Bear 方向列

        print(f"   -> 欄位映射：UL 使用「{ul_col}」，Bull/Bear 使用「{bb_col}」")

        df = df.rename(columns={ul_col: 'UL', bb_col: 'Bull/Bear'})
        # 去除重複欄位
        df = df.loc[:, ~df.columns.duplicated()]

        # 數值與文字處理
        for col in ['Strike Level', 'Total Issue Size', 'O/S (%)', 'Entitlement Ratio^']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '', regex=True), errors='coerce')

        df['UL'] = df['UL'].astype(str).str.strip()
        df['Bull/Bear'] = df['Bull/Bear'].astype(str).str.strip()

        print("   -> 映射後關鍵欄位預覽：")
        cols_debug = [c for c in [
            "CBBC Code", "Issuer", "UL", "Bull/Bear",
            "Strike Level", "Call Level",
            "Total Issue Size", "O/S (%)", "Entitlement Ratio^"
        ] if c in df.columns]
        if cols_debug:
            print(df[cols_debug].head(3).to_dict(orient="records"))
        else:
            print("      (無可用欄位預覽)")

        return df

    except Exception as e:
        print(f"[解析錯誤] {e}")
        raise


def get_hsi_cbbc(df):
    print("2. [篩選] 正在篩選 HSI...")

    # 寬鬆篩選 HSI
    mask = (
            df['UL'].str.upper().str.contains('HSI') |
            df['UL'].str.upper().str.contains('HANG SENG') |
            df['UL'].str.contains('恒生')
    )
    df_hsi = df[mask].copy()
    print(f"   -> 找到 {len(df_hsi)} 筆 HSI CBBC")

    # 如果還是 0，印出前 5 個 UL 讓我們檢查
    if df_hsi.empty:
        print(f"   !!! 警告：找不到 HSI。資料庫中的 UL 範例：{df['UL'].head(5).tolist()}")

    return df_hsi


def get_hsi_price():
    print("3. [市價] 獲取 HSI 價格...")
    for attempt in range(1, MAX_NETWORK_RETRIES + 1):
        try:
            hist = yf.Ticker("^HSI").history(period="5d")
            close = hist['Close'].iloc[-1]
            date = hist.index[-1].strftime('%Y%m%d')
            print(f"   -> {close:,.2f} ({date})")
            return round(float(close), 2), date
        except Exception as err:
            if attempt == MAX_NETWORK_RETRIES:
                print("   -> 無法連線，使用預設值 19200")
                break
            wait_s = min(5, 2 ** attempt)
            print(f"   -> 取價失敗 ({err})，{wait_s}s 後重試 ({attempt}/{MAX_NETWORK_RETRIES})")
            time.sleep(wait_s)
    return 19200.0, dt.datetime.now().strftime('%Y%m%d')


def build_price_detail(df_hsi, spot):
    columns = ['Price', 'Total_KO_$', 'Bull_KO_$', 'Bear_KO_$', 'Note']
    if df_hsi is None or df_hsi.empty:
        return pd.DataFrame(columns=columns)

    df = df_hsi.copy()
    for col in ['Strike Level', 'Entitlement Ratio^', 'Total Issue Size']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    if 'notional' not in df.columns:
        df['notional'] = (
                df['Entitlement Ratio^'] * df['Total Issue Size'] / 100 * spot)

    df['notional'] = df['notional'].fillna(0)
    df = df.dropna(subset=['Strike Level'])
    if df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        df.groupby(['Strike Level', 'Bull/Bear'])['notional']
            .sum()
            .unstack(fill_value=0)
    )

    for direction in ['Bull', 'Bear']:
        if direction not in summary.columns:
            summary[direction] = 0

    summary = summary[['Bull', 'Bear']]
    summary = summary.rename(columns={'Bull': 'Bull_KO_$', 'Bear': 'Bear_KO_$'})
    summary['Total_KO_$'] = summary['Bull_KO_$'] + summary['Bear_KO_$']

    price_detail = summary.reset_index().rename(columns={'Strike Level': 'Price'})
    price_detail = price_detail.sort_values('Price', ascending=False).reset_index(drop=True)

    # 将金额列转换为百万单位（与 CBBC KO 保持一致）
    for col in ['Total_KO_$', 'Bull_KO_$', 'Bear_KO_$']:
        price_detail[col] = price_detail[col] / 1e6
        price_detail[col] = price_detail[col].round(0)

    price_detail['Note'] = ''
    if not price_detail.empty:
        closest_idx = (price_detail['Price'] - spot).abs().idxmin()
        price_detail.loc[closest_idx, 'Note'] = 'Spot'

    return price_detail[['Price', 'Total_KO_$', 'Bull_KO_$', 'Bear_KO_$', 'Note']]


def format_value_for_html(value, fmt):
    try:
        if pd.isna(value):
            return ''
        if isinstance(value, str):
            return value
        return format(float(value), fmt)
    except Exception:
        return value


def apply_html_number_format(df, format_map):
    for col, fmt in format_map.items():
        if col in df.columns:
            df[col] = df[col].apply(lambda v: format_value_for_html(v, fmt))
    return df


def apply_excel_number_formats(ws, df, column_formats, has_index=True):
    if df is None or df.empty:
        return
    start_row = 2  # row 1 is header
    offset = 2 if has_index else 1  # account for index column when present
    max_row = len(df) + start_row
    for col, fmt in column_formats.items():
        if col not in df.columns:
            continue
        col_idx = df.columns.get_loc(col) + offset
        col_letter = get_column_letter(col_idx)
        for row in range(start_row, max_row):
            cell = ws[f"{col_letter}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt


def export_html(
        ladder: pd.DataFrame,
        price_detail: pd.DataFrame,
        spot: float,
        m_date: str,
        output_path: str,
        fx_rate: float | None = None):
    sheets = []
    if ladder is not None:
        sheets.append(('Ladder', ladder, True))
    if price_detail is not None and not price_detail.empty:
        sheets.append(('price_detail', price_detail, False))

    if not sheets:
        print("[警告] 沒有可用資料生成 HTML。")
        return

    default_tab = sheets[0][0]
    tab_buttons = []
    tab_contents = []

    columns_to_hide = ['Bear (M)', 'Bull (M)', 'BuI (M)', 'Bear Accu', 'Bull Accu', 'Bull/Bear']
    price_detail_hide = ['Bull', 'Bear', 'Bull_KO_$ (mio)', 'Bear_KO_$ (mio)', 'Bull/Bear', 'Note']

    for name, df, show_index in sheets:
        tab_id = f"tab-{name}"
        safe_name = name.lower().replace(' ', '_')
        table_id = f"table-{safe_name}"
        button = f'<button class="tab-link" data-tab="{name}">{name}</button>'
        tab_buttons.append(button)
        df_to_render = df.copy()
        render_index = show_index
        if name == 'Ladder':
            drop_cols = [c for c in columns_to_hide if c in df_to_render.columns]
            if drop_cols:
                df_to_render = df_to_render.drop(columns=drop_cols)
            df_to_render = df_to_render.reset_index().rename(columns={'index': 'range'})
            render_index = False
            df_to_render = apply_html_number_format(
                df_to_render,
                {
                    'Strike': ',.2f',
                    'CBBC KO (mio)': ',.0f'
                }
            )
        elif name == 'price_detail':
            drop_cols = [c for c in price_detail_hide if c in df_to_render.columns]
            if drop_cols:
                df_to_render = df_to_render.drop(columns=drop_cols)
            df_to_render.columns.name = None
            df_to_render.index.name = None
            df_to_render = df_to_render.reset_index(drop=True)
            render_index = False
            df_to_render = apply_html_number_format(
                df_to_render,
                {
                    'Price': ',.0f',
                    'Total_KO_$ (mio)': ',.0f'
                }
            )
        table_html = df_to_render.to_html(
            classes='dataframe',
            index=render_index,
            border=0,
            justify='center',
            table_id=table_id
        )
        content = (
            f'<div id="{tab_id}" class="tab-content">'
            f'{table_html}'
            '</div>'
        )
        tab_contents.append(content)

    tab_buttons_html = "\n".join(tab_buttons)
    tab_contents_html = "\n".join(tab_contents)
    html_template = f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8" />
<title>恆指牛熊打靶詳細列表 HSI CBBC Knock-Out Ladder</title>
<style>
    body {{
        font-family: Arial, "Microsoft JhengHei", "Microsoft YaHei", sans-serif;
        margin: 20px;
        background-color: #f7f7f7;
    }}
    h1 {{
        margin-bottom: 0;
    }}
    .meta {{
        margin-bottom: 20px;
        color: #555;
    }}
    .tabs {{
        margin-bottom: 20px;
    }}
    .tab-link {{
        border: none;
        padding: 10px 20px;
        margin-right: 10px;
        cursor: pointer;
        background-color: #e0e0e0;
        border-radius: 4px;
        font-size: 14px;
    }}
    .tab-link.active {{
        background-color: #007bff;
        color: #fff;
    }}
    .tab-content {{
        display: none;
        background-color: #fff;
        padding: 20px;
        border-radius: 6px;
        box-shadow: 0 2px 6px rgba(0,0,0,0.1);
    }}
    table.dataframe {{
        border-collapse: collapse;
        width: 100%;
        font-size: 12px;
    }}
    table.dataframe th, table.dataframe td {{
        border: 1px solid #ccc;
        padding: 6px;
        text-align: center;
        font-size: 14px;
        font-weight: bold;
        color: #000000;
    }}
    table.dataframe thead {{
        background-color: #f0f0f0;
    }}
</style>
</head>
<body>
    <h1>恆指牛熊打靶詳細列表 HSI CBBC Knock-Out Ladder</h1>
    <div class="meta">日期：{m_date} | 即時 HSI：{spot:.2f}{f" | 当前汇率：{fx_rate:.4f}" if fx_rate is not None else ""}</div>
    <div class="tabs">
        {tab_buttons_html}
    </div>
    {tab_contents_html}
<script>
    function showTab(name) {{
        var contents = document.querySelectorAll('.tab-content');
        var buttons = document.querySelectorAll('.tab-link');
        contents.forEach(function(content) {{
            content.style.display = 'none';
        }});
        buttons.forEach(function(btn) {{
            if (btn.getAttribute('data-tab') === name) {{
                btn.classList.add('active');
            }} else {{
                btn.classList.remove('active');
            }}
        }});
        var target = document.getElementById('tab-' + name);
        if (target) {{
            target.style.display = 'block';
        }}
    }}
    document.querySelectorAll('.tab-link').forEach(function(btn) {{
        btn.addEventListener('click', function() {{
            var tab = this.getAttribute('data-tab');
            showTab(tab);
        }});
    }});
    function interpolateColor(start, end, ratio) {{
        function hexToRgb(hex) {{
            var bigint = parseInt(hex.slice(1), 16);
            return {{
                r: (bigint >> 16) & 255,
                g: (bigint >> 8) & 255,
                b: bigint & 255
            }};
        }}
        function rgbToHex(r, g, b) {{
            return "#" + [r, g, b].map(function(x) {{
                var hex = Math.round(x).toString(16);
                return hex.length === 1 ? "0" + hex : hex;
            }}).join("");
        }}
        var startRGB = hexToRgb(start);
        var endRGB = hexToRgb(end);
        var r = startRGB.r + (endRGB.r - startRGB.r) * ratio;
        var g = startRGB.g + (endRGB.g - startRGB.g) * ratio;
        var b = startRGB.b + (endRGB.b - startRGB.b) * ratio;
        return rgbToHex(r, g, b);
    }}
    function applyGradientToColumn(tableId, columnName) {{
        var table = document.getElementById(tableId);
        if (!table) return;
        var headers = table.querySelectorAll('thead th');
        var colIndex = -1;
        headers.forEach(function(th, idx) {{
            if (th.textContent.trim() === columnName) {{
                colIndex = idx;
            }}
        }});
        if (colIndex === -1) return;
        var rows = table.querySelectorAll('tbody tr');
        var values = [];
        rows.forEach(function(row) {{
            var cell = row.cells[colIndex];
            if (!cell) return;
            var raw = cell.textContent.replace(/,/g, '');
            var value = parseFloat(raw);
            if (!isNaN(value)) {{
                cell.dataset.value = value;
                if (value > 0) {{
                    values.push(value);
                }}
            }}
        }});
        if (!values.length) return;
        values.sort(function(a, b) {{ return a - b; }});
        var denom = values.length - 1;
        var ratioMap = new Map();
        values.forEach(function(val, idx) {{
            if (!ratioMap.has(val)) {{
                var ratio = denom <= 0 ? 1 : idx / denom;
                ratioMap.set(val, ratio);
            }}
        }});
        rows.forEach(function(row) {{
            var cell = row.cells[colIndex];
            if (!cell) return;
            var value = parseFloat(cell.dataset.value);
            if (isNaN(value) || value <= 0) {{
                cell.style.backgroundColor = '#FFFFFF';
                cell.style.color = '#000000';
                return;
            }}
            var ratio = ratioMap.get(value) || 0;
            var steps = 10;
            ratio = Math.ceil(ratio * steps) / steps;
            ratio = Math.min(Math.max(ratio, 0.15), 1);
            var color = interpolateColor('#FFFFFF', '#00B050', ratio);
            cell.style.backgroundColor = color;
            cell.style.color = '#000000';
        }});
    }}
    function highlightBigValues(tableId, columnName, threshold) {{
        var table = document.getElementById(tableId);
        if (!table) return;
        var headers = table.querySelectorAll('thead th');
        var colIndex = -1;
        headers.forEach(function(th, idx) {{
            if (th.textContent.trim() === columnName) {{
                colIndex = idx;
            }}
        }});
        if (colIndex === -1) return;
        var rows = table.querySelectorAll('tbody tr');
        rows.forEach(function(row) {{
            var cell = row.cells[colIndex];
            if (!cell) return;
            var raw = cell.textContent.replace(/,/g, '');
            var value = parseFloat(raw);
            if (!isNaN(value) && value > threshold) {{
                cell.style.color = '#FF0000';
                cell.style.fontWeight = 'bold';
            }}
        }});
    }}
    showTab('{default_tab}');
    applyGradientToColumn('table-ladder', 'CBBC KO (mio)');
    applyGradientToColumn('table-price_detail', 'Total_KO_$ (mio)');
    highlightBigValues('table-ladder', 'CBBC KO (mio)', 1000000);
    highlightBigValues('table-price_detail', 'Total_KO_$ (mio)', 1000000);
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_template)
    print(f"[成功] HTML 檔案已產生：{output_path}")


def calculate_ladder(df_hsi, spot):
    if df_hsi.empty: return None
    print("4. [計算] 生成 Ladder...")

    # df_hsi['Entitlement Ratio^'] : true Total Issue Size
    # df_hsi['Total Issue Size']   : true O/S (%) (percentage)
    df_hsi['notional'] = (
                df_hsi['Entitlement Ratio^'] * df_hsi['Total Issue Size'] / 100 * spot).fillna(0)
    print("   -> notional describe (HSI only):")
    print(df_hsi['notional'].describe())
    bull = df_hsi[df_hsi['Bull/Bear'] == 'Bull']
    bear = df_hsi[df_hsi['Bull/Bear'] == 'Bear']
    print(f"   -> Bull count: {len(bull)}, Bear count: {len(bear)}")
    print(f"   -> Bull notional sum: {bull['notional'].sum():,.0f}")
    print(f"   -> Bear notional sum: {bear['notional'].sum():,.0f}")

    ladder = pd.DataFrame(index=strike_percent)
    ladder['Strike'] = [spot if x == 'Spot' else round(spot * (1 + int(x.strip('%')) / 100), 2) for x in ladder.index]
    ladder['Bear (M)'] = 0;
    ladder['Bull (M)'] = 0

    for i in range(1, 11):
        lower, upper = spot * (1 + (i - 1) / 100), spot * (1 + i / 100)
        m = (bear['Strike Level'] > lower) & (bear['Strike Level'] <= upper)
        ladder.loc[f'{i}%', 'Bear (M)'] = int(bear.loc[m, 'notional'].sum() / 1e6)

        upper_b, lower_b = spot * (1 - (i - 1) / 100), spot * (1 - i / 100)
        m_b = (bull['Strike Level'] < upper_b) & (bull['Strike Level'] >= lower_b)
        ladder.loc[f'-{i}%', 'Bull (M)'] = int(bull.loc[m_b, 'notional'].sum() / 1e6)
    # Debug: show Bear/Bull in 1%~3% buckets
    for i in range(1, 4):
        lower, upper = spot * (1 + (i - 1) / 100), spot * (1 + i / 100)
        mask_bear = (bear['Strike Level'] > lower) & (bear['Strike Level'] <= upper)
        bear_in_bucket = bear.loc[mask_bear, 'notional'].sum()

        upper_b, lower_b = spot * (1 - (i - 1) / 100), spot * (1 - i / 100)
        mask_bull = (bull['Strike Level'] < upper_b) & (bull['Strike Level'] >= lower_b)
        bull_in_bucket = bull.loc[mask_bull, 'notional'].sum()

        print(f"   -> Bucket {i}%: Bear notional sum = {bear_in_bucket:,.0f}, Bull notional sum = {bull_in_bucket:,.0f}")

    bear_idx = [f'{i}%' for i in range(1, 11) if f'{i}%' in ladder.index]
    if bear_idx: ladder.loc[bear_idx, 'Bear Accu'] = ladder.loc[bear_idx, 'Bear (M)'].cumsum()

    bull_idx = [f'-{i}%' for i in range(1, 11) if f'-{i}%' in ladder.index]
    if bull_idx: ladder.loc[bull_idx, 'Bull Accu'] = ladder.loc[bull_idx, 'Bull (M)'].cumsum()

    # --- New: CBBC KO column (total KO notional per bucket, in M) ---
    ladder['CBBC KO'] = ladder['Bear (M)'] + ladder['Bull (M)']

    # Optional: reorder columns so CBBC KO appears after Strike
    cols_order = ['Strike', 'CBBC KO', 'Bear (M)', 'Bull (M)', 'Bear Accu', 'Bull Accu']
    existing_cols = [c for c in cols_order if c in ladder.columns]
    extra_cols = [c for c in ladder.columns if c not in existing_cols]
    ladder = ladder[existing_cols + extra_cols]

    ladder.loc['Spot', ['Bear (M)', 'Bull (M)']] = ['SPOT', spot]
    ladder.loc['Spot', 'CBBC KO'] = 0

    return ladder.fillna(0)


def main():
    print("\n=== HSI CBBC Tool (V5 Final Debug) ===\n")
    try:
        df_all = download_and_parse_v5()
        df_hsi = get_hsi_cbbc(df_all)
        spot, m_date = get_hsi_price()
        rate, fx_ts = get_usdhkd_rate()
        if rate is None:
            print("   -> 無法獲取 USD/HKD 匯率，CBBC KO / Total_KO_$ 維持港幣金額")
        else:
            print(f"   -> 1 USD = {rate:.4f} HKD (資料時間: {fx_ts})")

        ladder = calculate_ladder(df_hsi, spot)
        price_detail = build_price_detail(df_hsi, spot)
        if rate is not None:
            # 对 Ladder 中的「CBBC KO」列做 HKD -> USD 换算
            if ladder is not None and 'CBBC KO' in ladder.columns:
                ladder['CBBC KO'] = ladder['CBBC KO'] / rate
            # 对 price_detail 中的金额列做 HKD -> USD 换算
            if (
                price_detail is not None
                and not price_detail.empty
            ):
                if 'Total_KO_$' in price_detail.columns:
                    price_detail['Total_KO_$'] = price_detail['Total_KO_$'] / rate
                if 'Bull_KO_$' in price_detail.columns:
                    price_detail['Bull_KO_$'] = price_detail['Bull_KO_$'] / rate
                if 'Bear_KO_$' in price_detail.columns:
                    price_detail['Bear_KO_$'] = price_detail['Bear_KO_$'] / rate

        if ladder is not None:
            # 为导出创建副本并重命名百万美元列
            ladder_export = ladder.copy()
            if 'CBBC KO' in ladder_export.columns:
                ladder_export = ladder_export.rename(columns={'CBBC KO': 'CBBC KO (mio)'})
            
            price_detail_export = None
            if price_detail is not None and not price_detail.empty:
                price_detail_export = price_detail.copy()
                rename_map = {}
                if 'Total_KO_$' in price_detail_export.columns:
                    rename_map['Total_KO_$'] = 'Total_KO_$ (mio)'
                if 'Bull_KO_$' in price_detail_export.columns:
                    rename_map['Bull_KO_$'] = 'Bull_KO_$ (mio)'
                if 'Bear_KO_$' in price_detail_export.columns:
                    rename_map['Bear_KO_$'] = 'Bear_KO_$ (mio)'
                if rename_map:
                    price_detail_export = price_detail_export.rename(columns=rename_map)
            
            out_file = spot_dir / f"HSI_CBBC_Ladder_{m_date}.xlsx"
            with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                ladder_export.to_excel(writer, sheet_name='Ladder')
                if 'CBBC KO (mio)' in ladder_export.columns:
                    ws = writer.sheets['Ladder']
                    start_row = 2
                    end_row = len(ladder_export) + 1
                    col_idx = ladder_export.columns.get_loc('CBBC KO (mio)') + 2  # +1 for index, +1 for 1-based
                    col_letter = get_column_letter(col_idx)
                    data_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                    rule = ColorScaleRule(
                        start_type='num', start_value=0, start_color='FFFFFF',
                        end_type='max', end_value=None, end_color='00B050'
                    )
                    ws.conditional_formatting.add(data_range, rule)
                    big_rule = CellIsRule(
                        operator='greaterThan',
                        formula=['1000000'],
                        font=Font(color='00FF0000', bold=True)
                    )
                    ws.conditional_formatting.add(data_range, big_rule)
                    apply_excel_number_formats(
                        ws,
                        ladder_export,
                        {
                            'Strike': '0.00',
                            'CBBC KO (mio)': '#,##0',
                            'Bear (M)': '#,##0',
                            'Bull (M)': '#,##0',
                            'Bear Accu': '#,##0',
                            'Bull Accu': '#,##0'
                        },
                        has_index=True
                    )
                if price_detail_export is not None:
                    price_detail_export.to_excel(writer, sheet_name='price_detail', index=False)
                    ws_detail = writer.sheets['price_detail']
                    apply_excel_number_formats(
                        ws_detail,
                        price_detail_export,
                        {
                            'Price': '#,##0',
                            'Total_KO_$ (mio)': '#,##0',
                            'Bull_KO_$ (mio)': '#,##0',
                            'Bear_KO_$ (mio)': '#,##0'
                        },
                        has_index=False
                    )
                    if 'Total_KO_$ (mio)' in price_detail_export.columns:
                        start_row = 2
                        end_row = len(price_detail_export) + 1
                        col_idx = price_detail_export.columns.get_loc('Total_KO_$ (mio)') + 1  # price_detail 沒有 index
                        col_letter = get_column_letter(col_idx)
                        data_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                        big_rule_detail = CellIsRule(
                            operator='greaterThan',
                            formula=['1000000'],
                            font=Font(color='00FF0000', bold=True)
                        )
                        ws_detail.conditional_formatting.add(data_range, big_rule_detail)
            print(f"\n[成功] 檔案已產生：{out_file}")
            print(f"Bear 1~3% 總街貨: {ladder.loc[['1%', '2%', '3%'], 'Bear (M)'].sum()} M")
            html_file = spot_dir / f"HSI_CBBC_Ladder_{m_date}.html"
            export_html(
                ladder=ladder_export,
                price_detail=price_detail_export,
                spot=spot,
                m_date=m_date,
                output_path=html_file,
                fx_rate=rate
            )
        else:
            print("\n[失敗] 沒有 HSI 資料，無法計算。")

    except Exception as e:
        print(f"\n[錯誤] {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()